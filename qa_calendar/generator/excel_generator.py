"""
QA 캘린더 Excel 생성기

Sheet 1: 연간 운영 캘린더 (기획점검 계획/수시, 대외기관 등)
Sheet 2: 상품 진행이슈 (방송상품, 현장QA, 암행)
"""
import openpyxl
from openpyxl.styles import (
    Font, Alignment, PatternFill, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter


MONTHS = ['1월', '2월', '3월', '4월', '5월', '6월',
          '7월', '8월', '9월', '10월', '11월', '12월']
MONTH_KEYS = ['01', '02', '03', '04', '05', '06',
              '07', '08', '09', '10', '11', '12']

SEASON_MAP = {
    '01': '새해', '02': '설 명절\n발렌타인데이', '03': '신학기\n화이트데이',
    '04': '봄(환절기)\n이사철', '05': '가정의 달', '06': '캠핑시즌\n물놀이',
    '07': '하절기', '08': '휴가철\n역시즌상품', '09': '추석 명절',
    '10': '가을(환절기)', '11': '김장철\n수능', '12': '크리스마스',
}

# 스타일 정의
HEADER_FILL = PatternFill('solid', fgColor='D9D9D9')
TITLE_FILL = PatternFill('solid', fgColor='BDD7EE')
ISSUE_FILL = PatternFill('solid', fgColor='FCE4D6')
RED_FONT = Font(name='맑은 고딕', bold=True, color='FF0000', size=10)
BOLD_FONT = Font(name='맑은 고딕', bold=True, size=10)
NORMAL_FONT = Font(name='맑은 고딕', size=9)
THIN_BORDER = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)


def _sanitize(text: str) -> str:
    """Excel에서 허용되지 않는 제어문자 제거"""
    import re
    if not isinstance(text, str):
        return text
    return re.sub(r'[\x00-\x08\x0b\x0c\x0e-\x1f]', '', text)


def _cell_style(ws, row, col, value='', bold=False, fill=None,
                font_size=9, halign='left', valign='top',
                wrap=True, color='000000'):
    if isinstance(value, str):
        value = _sanitize(value)
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = Font(name='맑은 고딕', bold=bold, size=font_size, color=color)
    cell.alignment = Alignment(
        horizontal=halign, vertical=valign,
        wrap_text=wrap
    )
    cell.border = THIN_BORDER
    if fill:
        cell.fill = fill
    return cell


def generate_calendar_excel(
    output_path: str,
    plan_items_by_month: dict,         # {month_key: {'상품': [...], '현장': [...], '모니터링': [...]}}
    unplanned_items_by_month: dict,    # {month_key: [inspection_name, ...]}
    issue_data: dict,                  # {month_key: {'방송상품': [...], '현장': [...], '암행': {...}}}
    other_data: dict = None,           # {row_name: {month_key: text}} for 대외기관 etc.
):
    """
    QA 캘린더 Excel 파일 생성

    Args:
        plan_items_by_month: 기획점검 계획 항목 (월별)
        unplanned_items_by_month: 수시점검 항목 (월별)
        issue_data: 상품QA 진행이슈 (월별)
        other_data: 대외기관, 품질교육 등 기타 행 데이터
    """
    wb = openpyxl.Workbook()

    _create_calendar_sheet(wb, plan_items_by_month, unplanned_items_by_month, other_data)
    _create_issues_sheet(wb, issue_data)

    wb.save(output_path)
    print(f"✅ 캘린더 저장 완료: {output_path}")


def _create_calendar_sheet(wb, plan_items_by_month, unplanned_items_by_month, other_data):
    ws = wb.active
    ws.title = '연간 운영 캘린더'

    # 제목
    ws.merge_cells('A1:P1')
    title_cell = ws['A1']
    title_cell.value = '26년 품질관리팀 운영 캘린더'
    title_cell.font = Font(name='맑은 고딕', bold=True, size=16)
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    title_cell.fill = TITLE_FILL
    ws.row_dimensions[1].height = 30

    # 컬럼 너비 설정
    ws.column_dimensions['A'].width = 8   # 구분
    ws.column_dimensions['B'].width = 6   # 계획/수시
    ws.column_dimensions['C'].width = 6   # 상품/현장/모니터링
    ws.column_dimensions['D'].width = 5   # 공란
    for i, col in enumerate(['E', 'F', 'G', 'H', 'I', 'J',
                              'K', 'L', 'M', 'N', 'O', 'P']):
        ws.column_dimensions[col].width = 22

    # 헤더행 (월)
    row = 2
    ws.row_dimensions[row].height = 18
    _cell_style(ws, row, 1, '구분', bold=True, fill=HEADER_FILL, halign='center', valign='center')
    ws.merge_cells(f'A{row}:D{row}')
    for i, month in enumerate(MONTHS):
        _cell_style(ws, row, 5 + i, month, bold=True, fill=HEADER_FILL,
                    halign='center', valign='center', font_size=10)

    # 시즌특성
    row = 3
    ws.row_dimensions[row].height = 32
    ws.merge_cells(f'A{row}:D{row}')
    _cell_style(ws, row, 1, '시즌특성', bold=True, fill=HEADER_FILL,
                halign='center', valign='center')
    for i, mk in enumerate(MONTH_KEYS):
        _cell_style(ws, row, 5 + i, SEASON_MAP.get(mk, ''),
                    halign='center', valign='center', font_size=9)

    # 기획점검 계획 - 상품
    start_product_row = 4
    product_rows = _write_plan_rows(
        ws, start_product_row, '기획점검', '계획', '상품',
        plan_items_by_month, sub_key='상품'
    )

    # 기획점검 계획 - 현장
    field_row = product_rows
    field_rows = _write_plan_rows(
        ws, field_row, '', '', '현장',
        plan_items_by_month, sub_key='현장'
    )

    # 기획점검 계획 - 모니터링
    monitor_row = field_rows
    monitor_rows = _write_plan_rows(
        ws, monitor_row, '', '', '모니터링',
        plan_items_by_month, sub_key='모니터링'
    )

    # 기획점검 수시
    suisi_row = monitor_rows
    suisi_end = _write_suisi_rows(ws, suisi_row, unplanned_items_by_month)

    # 대외기관 등 기타 행
    current_row = suisi_end
    if other_data:
        for row_name, month_data in other_data.items():
            current_row = _write_other_row(ws, current_row, row_name, month_data)

    # 기획점검 전체 왼쪽 그룹 레이블 병합
    _merge_group_label(ws, start_product_row, suisi_end - 1, 1, '기획점검')
    _merge_group_label(ws, start_product_row, monitor_rows - 1, 2, '계획')


def _write_plan_rows(ws, start_row, group_label, plan_label, sub_label,
                     plan_items_by_month, sub_key):
    """기획점검 계획 행 작성"""
    # 각 월별 항목 최대 수 계산
    max_items = max(
        (len(plan_items_by_month.get(mk, {}).get(sub_key, []))
         for mk in MONTH_KEYS),
        default=1
    )
    max_items = max(max_items, 3)

    row_height = max(20 * max_items, 60)

    # 레이블 셀
    ws.row_dimensions[start_row].height = row_height
    if sub_label:
        ws.merge_cells(f'D{start_row}:D{start_row}')
        _cell_style(ws, start_row, 4, sub_label, bold=True,
                    fill=HEADER_FILL, halign='center', valign='center', font_size=9)

    for i, mk in enumerate(MONTH_KEYS):
        items = plan_items_by_month.get(mk, {}).get(sub_key, [])
        text = '\n'.join(f'• {item}' for item in items)
        _cell_style(ws, start_row, 5 + i, text, valign='top', font_size=9)

    return start_row + 1


def _write_suisi_rows(ws, start_row, unplanned_by_month):
    """수시점검 행 작성"""
    max_items = max(
        (len(unplanned_by_month.get(mk, [])) for mk in MONTH_KEYS),
        default=1
    )
    max_items = max(max_items, 2)
    row_height = max(18 * max_items, 50)

    ws.row_dimensions[start_row].height = row_height
    ws.merge_cells(f'A{start_row}:B{start_row}')
    _cell_style(ws, start_row, 1, '수시', bold=True,
                fill=HEADER_FILL, halign='center', valign='center')
    ws.merge_cells(f'C{start_row}:D{start_row}')
    _cell_style(ws, start_row, 3, '', fill=HEADER_FILL)

    for i, mk in enumerate(MONTH_KEYS):
        items = unplanned_by_month.get(mk, [])
        text = '\n'.join(f'• {item}' for item in items)
        _cell_style(ws, start_row, 5 + i, text, valign='top', font_size=9)

    return start_row + 1


def _write_other_row(ws, row, label, month_data):
    """대외기관 등 기타 행 작성"""
    ws.row_dimensions[row].height = 60
    ws.merge_cells(f'A{row}:D{row}')
    _cell_style(ws, row, 1, label, bold=True,
                fill=HEADER_FILL, halign='center', valign='center', font_size=9)
    for i, mk in enumerate(MONTH_KEYS):
        text = month_data.get(mk, '')
        _cell_style(ws, row, 5 + i, text, valign='top', font_size=9)
    return row + 1


def _merge_group_label(ws, start_row, end_row, col, label):
    """그룹 레이블 셀 병합"""
    if end_row >= start_row:
        if end_row > start_row:
            ws.merge_cells(
                start_row=start_row, start_column=col,
                end_row=end_row, end_column=col
            )
        _cell_style(ws, start_row, col, label, bold=True,
                    fill=HEADER_FILL, halign='center', valign='center',
                    font_size=10)


def _create_issues_sheet(wb, issue_data):
    """상품 진행이슈 Sheet 생성"""
    ws = wb.create_sheet(title='상품 진행이슈')

    # 제목
    ws.merge_cells('A1:N1')
    ws['A1'].value = '26년 품질관리팀 상품 진행이슈'
    ws['A1'].font = Font(name='맑은 고딕', bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws['A1'].fill = ISSUE_FILL
    ws.row_dimensions[1].height = 28

    # 컬럼 너비
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 7
    for col in ['C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N']:
        ws.column_dimensions[col].width = 24

    # 헤더
    row = 2
    _cell_style(ws, row, 1, '구분', bold=True, fill=HEADER_FILL,
                halign='center', valign='center')
    _cell_style(ws, row, 2, '', fill=HEADER_FILL)
    for i, month in enumerate(MONTHS):
        _cell_style(ws, row, 3 + i, month, bold=True, fill=HEADER_FILL,
                    halign='center', valign='center', font_size=10)

    # 시즌특성
    row = 3
    ws.row_dimensions[row].height = 30
    ws.merge_cells(f'A{row}:B{row}')
    _cell_style(ws, row, 1, '시즌특성', bold=True, fill=HEADER_FILL,
                halign='center', valign='center')
    for i, mk in enumerate(MONTH_KEYS):
        _cell_style(ws, row, 3 + i, SEASON_MAP.get(mk, ''),
                    halign='center', valign='center', font_size=9)

    # 방송상품
    bs_start = 4
    bs_end = _write_issue_row(ws, bs_start, '방송\n상품', issue_data, '방송상품')

    # 현장
    field_start = bs_end
    field_end = _write_issue_row(ws, field_start, '현장', issue_data, '현장QA')

    # 암행
    secret_start = field_end
    secret_end = _write_암행_row(ws, secret_start, issue_data)

    # 왼쪽 그룹 레이블 병합
    label_start = bs_start
    label_end = secret_end - 1
    if label_end >= label_start:
        ws.merge_cells(
            start_row=label_start, start_column=1,
            end_row=label_end, end_column=1
        )
        _cell_style(ws, label_start, 1, '상\n품\nQ\nA\n진\n행\n이\n슈',
                    bold=True, fill=ISSUE_FILL, halign='center',
                    valign='center', font_size=11)


def _write_issue_row(ws, row, label, issue_data, category_key):
    """방송상품/현장 이슈 행 작성"""
    # 각 월별 최대 항목 수
    max_items = 1
    for mk in MONTH_KEYS:
        items = issue_data.get(mk, {}).get(category_key, [])
        max_items = max(max_items, len(items))

    row_height = max(80 * max_items, 80)
    ws.row_dimensions[row].height = row_height

    _cell_style(ws, row, 2, label, bold=True, fill=HEADER_FILL,
                halign='center', valign='center', font_size=9)

    for i, mk in enumerate(MONTH_KEYS):
        items = issue_data.get(mk, {}).get(category_key, [])
        text = _format_qa_issues(items, category_key)
        _cell_style(ws, row, 3 + i, text, valign='top', font_size=8)

    return row + 1


def _write_암행_row(ws, row, issue_data):
    """암행점검 이슈 행 작성"""
    ws.row_dimensions[row].height = 100
    _cell_style(ws, row, 2, '암행', bold=True, fill=HEADER_FILL,
                halign='center', valign='center', font_size=9)

    for i, mk in enumerate(MONTH_KEYS):
        암행 = issue_data.get(mk, {}).get('암행', {})
        text = _format_암행(암행)
        _cell_style(ws, row, 3 + i, text, valign='top', font_size=8)

    return row + 1


def _format_qa_issues(issues: list, category: str) -> str:
    """QAIssue 리스트를 셀 텍스트로 변환"""
    if not issues:
        return ''

    parts = []
    for issue in issues:
        if category == '방송상품':
            result_tag = f'[상품QA – {issue.result}]' if issue.result else '[상품QA]'
            line = f'{result_tag}\n{issue.product_name}'
            if issue.md_name:
                line += f'\n({issue.md_name})'
            if issue.issue:
                # 이슈 첫 줄만
                first_issue = issue.issue.split('\n')[0].strip()[:80]
                line += f'\n- {first_issue}'
            if issue.measure:
                line += f'\n→ {issue.measure[:60]}'
        elif category == '현장QA':
            result_tag = f'[현장QA – {issue.result}]' if issue.result else '[현장QA]'
            line = f'{result_tag}\n{issue.product_name}'
            if issue.md_name:
                line += f'\n({issue.md_name})'
            if issue.issue:
                first_issue = issue.issue.split('\n')[0].strip()[:80]
                line += f'\n- {first_issue}'
        else:
            line = issue.product_name
        parts.append(line)

    return '\n\n'.join(parts)


def _format_암행(암행_dict: dict) -> str:
    """암행점검 부적합 데이터를 유형별 텍스트로 변환"""
    if not 암행_dict:
        return ''

    parts = []
    for 유형, products in sorted(암행_dict.items()):
        count = len(products)
        lines = [f'{유형}({count}건)']
        for p in products[:8]:  # 최대 8개
            lines.append(f'- {p[:30]}')
        parts.append('\n'.join(lines))

    return '\n\n'.join(parts)
